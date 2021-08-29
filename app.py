# Made By SepehrRS.
from openpyxl import load_workbook
import datetime
import openpyxl
import configdata
import matplotlib.pyplot as plt
from itertools import chain, repeat, takewhile
defaultdatabase = "bookdatabase.xlsx"


def addBook():
    print("Okay. Im Adding New books To Your Database.")
    print("If You Want To Add More Than One Book , Seperate them With a //")
    print("Writer Name/Names ?")
    writer = input("> ")
    writer = writer.split("//") if '//' in writer else writer
    print("Book Name/Names ?")
    bookName = input("> ")
    bookName = bookName.split("//") if '//' in bookName else bookName
    print("Genre/Genres ? ")
    genre = input("> ")
    genre = genre.split("//") if '//' in genre else genre

    print("Personal Review ? write NO to continue without pesrsonal review")
    review = input("> ")
    if review == "NO":
        review = ''
    else:
        review = review.split("//") if '//' in review else review

    if type(writer) == list:
        # If The user gives more than two data.
        print(f"Adding Book To {configdata.DATABASEPATH}")
        writeToDatabase([writer,bookName,genre,review])

    elif type(writer) == str:
        print(f"Adding Book To {configdata.DATABASEPATH}")
        writeToDatabase([writer,bookName,genre,review])




def addBookReview():
    print("Ok. Adding book review to a existing book.")
    print("Genre?")
    genre = input("> ")
    print("Writer?")
    writer = input("> ")
    print("Book name?")
    bookname = input("> ")
    print("Review ? ")
    review = input("> ")
    listOfBooks = []
    listOfData = [bookname,genre]
    wb = load_workbook(filename=configdata.DATABASEPATH)
    ws = wb['Books']
    for data in ws.iter_cols():
        for values in data:
            if values.value and values.value.strip() == writer:
                listOfBooks.append(values)

            
    if not listOfBooks:
        print("Writer Not Found.")
        takeInput()


    for element in listOfBooks:
        tempList = []
        for num in range(1,3):
            if ws.cell(row=element.row+num,column=element.column+num) and \
                ws.cell(row=element.row+num,column=element.column+num) == listOfData[num-1]:
                tempList.append(True)

        if all(tempList) == True:
            wb = load_workbook(filename=configdata.DATABASEPATH)
            ws = wb['Books']
            ws.cell(row=element.row,column=4,value=review)
            wb.save(configdata.DATABASEPATH)
            wb.close()
            print("Done.")
            takeInput()

        else:
            continue

    else:
        print("Book not found.")
        takeInput()

    


    

def showBookList():
    print("Filters?")
    print("1- By Genre 2- By Writer name 3- By Bookname 4-Show All")
    nums = ['1','2','3','4']
    prompts = chain(["> "], repeat("Invalid Choice , Try Again > "))
    replies = map(input, prompts)
    valid_response = next(filter(nums.__contains__, replies))
    if valid_response == '1':
        print("Genre ?")
        genre = input("> ")
        wb = load_workbook(filename=configdata.DATABASEPATH)
        ws = wb['Books']
        for data in ws.iter_cols():
            for values in data:
                if values.value and genre.lower() in values.value.strip().lower():
                    writer = ws.cell(row=values.row,column=1).value 
                    bookname = ws.cell(row=values.row,column=2).value
                    review = ws.cell(row=values.row,column=4).value
                    print(f"Writer Name : {writer}, Book Name : {bookname}, Genre : {genre}, Review : {review}")
        wb.close()
        takeInput()
    if valid_response == '2':
        print("Writer ?")
        writer = input("> ")
        wb = load_workbook(filename=configdata.DATABASEPATH)
        ws = wb['Books']
        for data in ws.iter_cols():
            for values in data:
                if values.value and writer.lower() in values.value.strip().lower():
                    genre = ws.cell(row=values.row,column=3).value 
                    bookname = ws.cell(row=values.row,column=2).value
                    review = ws.cell(row=values.row,column=4).value
                    print(f"Writer Name : {writer}, Book Name : {bookname}, Genre : {genre}, Review : {review}")
        wb.close()
        takeInput()
    if valid_response == '3':
        print("By Book Name?")
        bookname = input("> ")
        wb = load_workbook(filename=configdata.DATABASEPATH)
        ws = wb['Books']
        for data in ws.iter_cols():
            for values in data:
                if values.value and bookname.lower() in values.value.strip().lower():
                    writer = ws.cell(row=values.row,column=1).value 
                    genre = ws.cell(row=values.row,column=3).value
                    review = ws.cell(row=values.row,column=4).value
                    print(f"Writer Name : {writer}, Book Name : {bookname}, Genre : {genre}, Review : {review}")

        wb.close()
        takeInput()
    else:
        print("Showing all data.")
        wb = load_workbook(filename=configdata.DATABASEPATH)
        ws = wb['Books']
        for data in ws.iter_cols():
            for values in data:
                if values.value:
                    writer = ws.cell(row=values.row,column=1).value 
                    bookname = ws.cell(row=values.row,column=2).value
                    genre = ws.cell(row=values.row,column=3).value
                    review = ws.cell(row=values.row,column=4).value
                    print(f"Writer Name : {writer}, Book Name : {bookname}, Genre : {genre}, Review : {review}")
        

    
def showChartByGenre():
    wb = load_workbook(filename=configdata.DATABASEPATH)
    ws = wb['Books']
    genres = []
    value_per_genre = {}
    for data in ws.iter_cols(min_col=3,max_col=3):
        for genre in data:
            if genre.value:
                genres.append(genre.value)

    for genre in genres:
        if genre in value_per_genre.keys():
            continue

        else:
            value_per_genre[genre] = genres.count(genre)

    GenreNames = list(i for i in value_per_genre.keys())
    GenreValues = list(i for i in value_per_genre.values())
    plt.bar(GenreNames,GenreValues)
    plt.title('Genres')
    plt.show()
    takeInput()
    


def showChartByWriterName():
    wb = load_workbook(filename=configdata.DATABASEPATH)
    ws = wb['Books']
    writers = []
    value_per_writer = {}
    for data in ws.iter_cols(min_col=1,max_col=1):
        for writer in data:
            if writer.value:
                writers.append(writer.value)

    for writer in writers:
        if writer in value_per_writer.keys():
            continue

        else:
            value_per_writer[writer] = writers.count(writer)

    WriterNames = list(i for i in value_per_writer.keys())
    WriterValues = list(i for i in value_per_writer.values())
    plt.bar(WriterNames,WriterValues)
    plt.title('Writers')
    plt.show()
    takeInput()
                

def showChartInput():
    print("Filters?")
    print("1- By Genre 2- By Writer name")
    validChoices = ['1','2']
    prompts = chain(["> "],repeat("Invalid Choice. Try Again. > "))
    responses = map(input,prompts)
    userInput = next(filter(validChoices.__contains__,responses))
    if userInput == "1":
        showChartByGenre()

    if userInput == "2":
        showChartByWriterName()


def writeToDatabase(contents):
    try:
        
        wb = load_workbook(filename=configdata.DATABASEPATH)
        ws = wb['Books']
        if type(contents[0]) == list:

            for col in ws.iter_cols():
                # to check if the book name already existd
                for value in col:
                    if value.value and value.value.strip() in contents[2]:
                        print("Book Already Exists") 

                       
            for i in range(0,len(contents[0])):
                templist = []
                for j in range(0,len(contents)):
                    templist.append(contents[j][i])
                
                ws.append(templist)


            wb.save(configdata.DATABASEPATH)
            wb.close()
            takeInput()
        else:
            ws.append(contents)
            wb.save(configdata.DATABASEPATH)
            wb.close()
            takeInput()
    except PermissionError:
        print("Please Close Excel File.")
    

def getTime():
    currentTime = datetime.datetime.now()
    if currentTime.hour < 12:
        return 'Good morning'
    elif 12 <= currentTime.hour < 18:
        return 'Good afternoon'
    else:
        return 'Good evening'



def takeInput():
    """ Take user choices"""
    print("*"*10)
    time = getTime()
    print(f"Hi. {time}. Im Your Book Assistant, I Hope You're Doing Well.")
    print("""
    Options :
        1- Add a new Book
        2- Add Book Review To Existing Book
        3- Show Book List
        4- Book Chart
        """)
    correctNumbers = ['1','2','3','4']
    prompts = chain(["> "],repeat("Invalid Choice, Try again > "))
    responses = map(input,prompts)
    valid_responses = next(filter(correctNumbers.__contains__,responses))
    if valid_responses == "1":
        addBook()
    elif valid_responses == "2":
        addBookReview()

    elif valid_responses == "3":
        showBookList()

    elif valid_responses == "4":
        showChartInput()

    
takeInput()